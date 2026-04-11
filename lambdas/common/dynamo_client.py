"""
DynamoDB helpers shared by all Lambda functions.
Implements hybrid store-scoped + global fallback mapping lookups.

Confidence thresholds (v4 agreed):
  >= 0.92  silent, confident mapping
  0.75–0.91 flag review, tentative mapping
  0.50–0.74 flag review, no mapping written (too uncertain to learn from)
  < 0.50   unknown, no mapping written
"""

import os
import logging
from datetime import datetime, timezone
from typing import Optional

import boto3
from boto3.dynamodb.conditions import Attr, Key
from decimal import Decimal

from models import (
    THRESHOLD_SILENT, THRESHOLD_REVIEW, THRESHOLD_GUESS,
    TRUST_PROMOTE_CONFIDENT, TRUST_PROMOTE_TRUSTED, MAX_PROCESS_COUNT,
    TrustLevel
)

logger = logging.getLogger(__name__)

RECEIPTS_TABLE   = os.environ["RECEIPTS_TABLE"]
ITEMS_TABLE      = os.environ["ITEMS_TABLE"]
MAPPINGS_TABLE   = os.environ["MAPPINGS_TABLE"]
CATEGORIES_TABLE = os.environ["CATEGORIES_TABLE"]
RETAILERS_TABLE  = os.environ["RETAILERS_TABLE"]
ALERT_TOPIC_ARN  = os.environ.get("ALERT_TOPIC_ARN", "")  # SNS topic for cost alerts
ITEM_TYPES_TABLE = os.environ.get("ITEM_TYPES_TABLE", "")
IMAGES_BUCKET    = os.environ.get("IMAGES_BUCKET", "")
STARLING_TABLE   = os.environ.get("STARLING_TABLE", "")

GLOBAL_STORE = "global"

_db  = None
_sns = None


def db():
    global _db
    if _db is None:
        _db = boto3.resource("dynamodb")
    return _db


def sns():
    global _sns
    if _sns is None:
        _sns = boto3.client("sns")
    return _sns


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _make_key(store_id: str, normalized_name: str) -> str:
    return f"{store_id}#{normalized_name}"


# =============================================================================
# RECEIPTS
# =============================================================================

def save_receipt(receipt: dict):
    db().Table(RECEIPTS_TABLE).put_item(Item=receipt)


def get_receipt(receipt_id: str) -> Optional[dict]:
    resp = db().Table(RECEIPTS_TABLE).get_item(Key={"receipt_id": receipt_id})
    item = resp.get("Item")
    return _decimals_to_float(item) if item else None


def update_receipt(receipt_id: str, **fields):
    """Update arbitrary fields on a receipt record."""
    table = db().Table(RECEIPTS_TABLE)
    fields["updated_at"] = _now()

    set_parts  = [f"#f{i} = :v{i}" for i, k in enumerate(fields)]
    attr_names = {f"#f{i}": k for i, k in enumerate(fields)}
    attr_vals  = {f":v{i}": v for i, (k, v) in enumerate(fields.items())}

    table.update_item(
        Key={"receipt_id": receipt_id},
        UpdateExpression="SET " + ", ".join(set_parts),
        ExpressionAttributeNames=attr_names,
        ExpressionAttributeValues=attr_vals
    )


def list_receipts(user_id: str = "default",
                  date_from: str = None,
                  date_to:   str = None) -> list[dict]:
    table  = db().Table(RECEIPTS_TABLE)
    kwargs = {
        "IndexName": "UserDateIndex",
        "KeyConditionExpression": Key("user_id").eq(user_id),
        "ScanIndexForward": False
    }
    if date_from and date_to:
        kwargs["KeyConditionExpression"] &= Key("receipt_date").between(date_from, date_to)

    items = []
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    return [_decimals_to_float(i) for i in items]


def get_receipts_summary(user_id: str = "default") -> dict:
    """Lightweight dashboard summary — counts + latest receipt. Uses ProjectionExpression
    to minimise data transfer (no large fields like s3_key, error_message, etc.)."""
    table = db().Table(RECEIPTS_TABLE)
    kwargs = {
        "IndexName": "UserDateIndex",
        "KeyConditionExpression": Key("user_id").eq(user_id),
        "ScanIndexForward": False,
        "ProjectionExpression": "#rid, #st, #ta, #ex, #rn, #ri, #rd, #ic",
        "ExpressionAttributeNames": {
            "#rid": "receipt_id",
            "#st":  "status",
            "#ta":  "total_amount",
            "#ex":  "exported_to_excel",
            "#rn":  "retailer_name",
            "#ri":  "retailer_id",
            "#rd":  "receipt_date",
            "#ic":  "item_count",
        }
    }
    items = []
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    items = [_decimals_to_float(i) for i in items]
    total        = len(items)
    needs_review = sum(1 for i in items if i.get("status") == "needs_review")
    exported     = sum(1 for i in items if i.get("exported_to_excel"))
    spend        = 0.0
    for i in items:
        try:
            spend += float(str(i.get("total_amount") or "0").replace("£", "").replace(",", ""))
        except (ValueError, TypeError):
            pass
    latest = items[0] if items else None  # most recent (ScanIndexForward=False)
    return {
        "total":        total,
        "needs_review": needs_review,
        "exported":     exported,
        "total_spend":  f"£{spend:.2f}",
        "latest":       latest,
    }


def list_receipts_page(user_id: str = "default",
                       before: str = None,
                       days: int = 3) -> dict:
    """Return receipts in a date window for paginated list view.
    `before` is exclusive upper bound (YYYY-MM-DD); defaults to tomorrow so today is included.
    Returns receipts + next_cursor (start of this window) + has_more flag."""
    from datetime import date, timedelta
    today = date.today()
    end   = date.fromisoformat(before) if before else today + timedelta(days=1)
    start = end - timedelta(days=days)
    table = db().Table(RECEIPTS_TABLE)
    kwargs = {
        "IndexName": "UserDateIndex",
        "KeyConditionExpression": (
            Key("user_id").eq(user_id) &
            Key("receipt_date").between(start.isoformat(),
                                        (end - timedelta(days=1)).isoformat())
        ),
        "ScanIndexForward": False,
        "ProjectionExpression": "#rid, #st, #ta, #ex, #rn, #ri, #rd, #ic",
        "ExpressionAttributeNames": {
            "#rid": "receipt_id",
            "#st":  "status",
            "#ta":  "total_amount",
            "#ex":  "exported_to_excel",
            "#rn":  "retailer_name",
            "#ri":  "retailer_id",
            "#rd":  "receipt_date",
            "#ic":  "item_count",
        }
    }
    items = []
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]

    # Also scan for needs_review receipts that may have a missing/malformed date
    # (they won't appear in the date-range query above)
    seen_ids = {i["receipt_id"]["S"] if isinstance(i.get("receipt_id"), dict) else i.get("receipt_id") for i in items}
    review_resp = table.scan(
        FilterExpression=Attr("user_id").eq(user_id) & Attr("status").eq("needs_review"),
        ProjectionExpression="#rid, #st, #ta, #ex, #rn, #ri, #rd, #ic",
        ExpressionAttributeNames={
            "#rid": "receipt_id",
            "#st":  "status",
            "#ta":  "total_amount",
            "#ex":  "exported_to_excel",
            "#rn":  "retailer_name",
            "#ri":  "retailer_id",
            "#rd":  "receipt_date",
            "#ic":  "item_count",
        }
    )
    for r in review_resp.get("Items", []):
        rid = r.get("receipt_id") if not isinstance(r.get("receipt_id"), dict) else r["receipt_id"]["S"]
        if rid not in seen_ids:
            items.append(r)

    # Check if there are any receipts older than this window (one cheap Limit=1 query)
    check_kwargs = {
        "IndexName": "UserDateIndex",
        "KeyConditionExpression": (
            Key("user_id").eq(user_id) &
            Key("receipt_date").lt(start.isoformat())
        ),
        "Limit": 1,
        "ScanIndexForward": False,
        "ProjectionExpression": "#rid",
        "ExpressionAttributeNames": {"#rid": "receipt_id"},
    }
    has_more = len(table.query(**check_kwargs).get("Items", [])) > 0
    return {
        "receipts":    [_decimals_to_float(i) for i in items],
        "next_cursor": start.isoformat(),
        "has_more":    has_more,
    }


def delete_receipt(receipt_id: str):
    db().Table(RECEIPTS_TABLE).delete_item(Key={"receipt_id": receipt_id})


def check_and_increment_process_count(receipt_id: str) -> bool:
    """
    Cost safeguard: atomically check and increment process_count.
    Returns True if processing is allowed (count was < MAX_PROCESS_COUNT).
    Returns False if limit reached — caller must stop and set status=failed.

    Uses a conditional update so this is safe even if two Lambda instances
    run simultaneously (which shouldn't happen with concurrency=2, but safe anyway).
    """
    table = db().Table(RECEIPTS_TABLE)
    try:
        table.update_item(
            Key={"receipt_id": receipt_id},
            UpdateExpression="SET process_count = if_not_exists(process_count, :zero) + :inc, "
                             "updated_at = :ts",
            ConditionExpression="attribute_not_exists(process_count) OR process_count < :max",
            ExpressionAttributeValues={
                ":zero": 0,
                ":inc":  1,
                ":max":  MAX_PROCESS_COUNT,
                ":ts":   _now()
            }
        )
        return True
    except db().meta.client.exceptions.ConditionalCheckFailedException:
        # process_count >= MAX_PROCESS_COUNT
        logger.error("COST SAFEGUARD: receipt %s hit MAX_PROCESS_COUNT=%d", receipt_id, MAX_PROCESS_COUNT)
        _send_alert(f"Receipt {receipt_id} hit max reprocess limit ({MAX_PROCESS_COUNT}). "
                    f"Processing stopped. Check CloudWatch logs for details.")
        return False


def _send_alert(message: str):
    """Send SNS alert email. Silently skips if topic not configured."""
    if not ALERT_TOPIC_ARN:
        logger.warning("ALERT (no SNS topic configured): %s", message)
        return
    try:
        sns().publish(
            TopicArn=ALERT_TOPIC_ARN,
            Subject="Grocery Scanner Alert",
            Message=message
        )
    except Exception as e:
        logger.error("Failed to send SNS alert: %s", e)


# =============================================================================
# ITEMS
# =============================================================================

def _floats_to_decimal(obj):
    """Recursively convert float values to Decimal for DynamoDB compatibility."""
    if isinstance(obj, float):
        return Decimal(str(obj))
    elif isinstance(obj, dict):
        return {k: _floats_to_decimal(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_floats_to_decimal(v) for v in obj]
    return obj


def save_items(receipt_id: str, items: list[dict]):
    table = db().Table(ITEMS_TABLE)
    with table.batch_writer() as batch:
        for item in items:
            clean = _floats_to_decimal(item)
            batch.put_item(Item={"receipt_id": receipt_id, **clean})


def _decimals_to_float(obj):
    """Recursively convert Decimal values back to float for JSON serialisation."""
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {k: _decimals_to_float(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_decimals_to_float(v) for v in obj]
    return obj


def get_items(receipt_id: str) -> list[dict]:
    resp = db().Table(ITEMS_TABLE).query(
        KeyConditionExpression=Key("receipt_id").eq(receipt_id)
    )
    items = resp.get("Items", [])
    items = [_decimals_to_float(i) for i in items]
    return sorted(items, key=lambda x: x.get("item_seq", ""))


def update_item_category(receipt_id: str, item_seq: str, category: str,
                         item_type_id: str = None, record_correction: bool = False):
    expr = "SET category = :cat, confirmed = :yes, match_source = :src, updated_at = :ts"
    vals = {
        ":cat": category,
        ":yes": True,
        ":src": "manual",
        ":ts":  _now()
    }
    if item_type_id:
        expr += ", item_type_id = :iti"
        vals[":iti"] = item_type_id
    if record_correction:
        expr += ", corrected_at = :cor"
        vals[":cor"] = _now()
    db().Table(ITEMS_TABLE).update_item(
        Key={"receipt_id": receipt_id, "item_seq": item_seq},
        UpdateExpression=expr,
        ExpressionAttributeValues=vals
    )


def update_item_price(receipt_id: str, item_seq: str, price: str):
    """Update the price of a single line item."""
    db().Table(ITEMS_TABLE).update_item(
        Key={"receipt_id": receipt_id, "item_seq": item_seq},
        UpdateExpression="SET price = :p",
        ExpressionAttributeValues={":p": price},
    )


def delete_item(receipt_id: str, item_seq: str):
    """Delete a single line item."""
    db().Table(ITEMS_TABLE).delete_item(
        Key={"receipt_id": receipt_id, "item_seq": item_seq}
    )


def replace_item(receipt_id: str, item: dict):
    """Replace or create a single item record (used for split operation)."""
    clean = _floats_to_decimal(item)
    db().Table(ITEMS_TABLE).put_item(Item={"receipt_id": receipt_id, **clean})


def delete_items(receipt_id: str):
    """Delete all items for a receipt — used before reprocessing."""
    items = get_items(receipt_id)
    table = db().Table(ITEMS_TABLE)
    with table.batch_writer() as batch:
        for item in items:
            batch.delete_item(Key={
                "receipt_id": receipt_id,
                "item_seq":   item["item_seq"]
            })


# =============================================================================
# MAPPINGS — hybrid store-scoped + global fallback
#
# Key format:  "{store_id}#{normalized_name}"
#   "tesco#smoked salmon"   → store-scoped
#   "global#smoked salmon"  → global fallback (seeded)
#
# One BatchGetItem call checks both store-scoped and global simultaneously.
# Store-scoped result takes priority over global if both exist.
#
# After any fuzzy/partial match succeeds, the result is written back as a
# store-scoped mapping so future scans hit Layer 1 (exact) instead.
#
# Only mappings with confidence >= THRESHOLD_REVIEW (0.75) are persisted.
# Below that threshold the match is too uncertain to learn from.
# =============================================================================

def get_mapping_hybrid(store_id: str, normalized_name: str) -> Optional[dict]:
    """
    Single BatchGetItem that checks store-scoped AND global simultaneously.
    Returns the store-scoped mapping if found, else global, else None.
    """
    keys_to_fetch = [{"mapping_key": _make_key(GLOBAL_STORE, normalized_name)}]

    if store_id and store_id not in (GLOBAL_STORE, "unknown"):
        keys_to_fetch.insert(0, {"mapping_key": _make_key(store_id, normalized_name)})

    resp  = db().batch_get_item(RequestItems={MAPPINGS_TABLE: {"Keys": keys_to_fetch}})
    items = resp.get("Responses", {}).get(MAPPINGS_TABLE, [])

    store_hit  = next((i for i in items if i.get("store_id") == store_id),       None)
    global_hit = next((i for i in items if i.get("store_id") == GLOBAL_STORE),   None)

    return store_hit or global_hit


def promote_mapping_if_ready(mapping_key: str):
    """
    Increment match_count and promote trust level if thresholds are met.
    tentative →(3×)→ confident →(5×)→ trusted
    Trusted mappings are never demoted.
    """
    table = db().Table(MAPPINGS_TABLE)

    resp = table.update_item(
        Key={"mapping_key": mapping_key},
        UpdateExpression=(
            "SET match_count = if_not_exists(match_count, :zero) + :inc, "
            "last_seen = :ts"
        ),
        ExpressionAttributeValues={":zero": 0, ":inc": 1, ":ts": _now()},
        ReturnValues="ALL_NEW"
    )
    updated   = resp.get("Attributes", {})
    count     = int(updated.get("match_count", 0))
    trust     = updated.get("trust", "tentative")

    new_trust = None
    if trust == "tentative" and count >= TRUST_PROMOTE_CONFIDENT:
        new_trust = TrustLevel.CONFIDENT.value
    elif trust == "confident" and count >= TRUST_PROMOTE_TRUSTED:
        new_trust = TrustLevel.TRUSTED.value

    if new_trust:
        table.update_item(
            Key={"mapping_key": mapping_key},
            UpdateExpression="SET trust = :t",
            ExpressionAttributeValues={":t": new_trust}
        )
        logger.info("Promoted mapping %s → %s", mapping_key, new_trust)


def write_learned_mapping(store_id: str, normalized_name: str,
                           item_type_id: str, category: str, confidence: float,
                           source: str) -> None:
    """
    Persist a fuzzy match result to the store-scoped table so future identical
    OCR strings hit Layer 1 (exact) rather than going through fuzzy matching again.

    Stores item_type_id for the two-hop chain; category kept for backwards compat.

    Rules:
    - confidence >= 0.92 → write as confident
    - confidence >= 0.75 → write as tentative
    - confidence <  0.75 → do NOT write (too uncertain to learn from)
    - never overwrite a trusted mapping with a fuzzy result
    - never write if store is unknown (don't know which template to update)
    """
    if confidence < THRESHOLD_REVIEW:
        return  # below 0.75 — don't learn from this
    if not store_id or store_id == "unknown":
        return

    trust       = TrustLevel.CONFIDENT.value if confidence >= THRESHOLD_SILENT else TrustLevel.TENTATIVE.value
    mapping_key = _make_key(store_id, normalized_name)
    now         = _now()

    existing = db().Table(MAPPINGS_TABLE).get_item(
        Key={"mapping_key": mapping_key}
    ).get("Item")

    if existing and existing.get("trust") == TrustLevel.TRUSTED.value:
        return  # never overwrite trusted with a fuzzy result

    db().Table(MAPPINGS_TABLE).put_item(Item={
        "mapping_key":     mapping_key,
        "store_id":        store_id,
        "normalized_name": normalized_name,
        "item_type_id":    item_type_id,
        "category":        category,        # kept for backwards compat
        "confidence":      str(round(confidence, 4)),
        "match_count":     1,
        "trust":           trust,
        "source":          source,
        "created_at":      existing.get("created_at", now) if existing else now,
        "last_seen":       now
    })
    logger.info("Learned mapping %s → %s (%.0f%%, %s)", mapping_key, category, confidence * 100, trust)


def save_correction(store_id: str, normalized_name: str, item_type_id: str, category: str) -> dict:
    """
    Save a user correction back to the mappings table.

    Behaviour by existing trust level:
    - trusted + same category  → just increment match_count, no change
    - trusted + different category → save conflict flag, return conflict info
    - confident or tentative → overwrite with manual correction (instantly trusted)
    - missing → create new trusted mapping

    Returns dict with key 'conflict': True/False
    """
    effective_store = store_id if (store_id and store_id != "unknown") else GLOBAL_STORE
    mapping_key     = _make_key(effective_store, normalized_name)
    table           = db().Table(MAPPINGS_TABLE)
    now             = _now()

    existing = table.get_item(Key={"mapping_key": mapping_key}).get("Item")

    if existing and existing.get("trust") == TrustLevel.TRUSTED.value:
        if existing.get("item_type_id", existing.get("category")) == item_type_id:
            # User confirmed — just increment count
            promote_mapping_if_ready(mapping_key)
            return {"conflict": False}
        else:
            # Conflict — save it but don't overwrite
            table.update_item(
                Key={"mapping_key": mapping_key},
                UpdateExpression="SET conflict_category = :c, conflict_item_type = :it, conflict_at = :ts",
                ExpressionAttributeValues={":c": category, ":it": item_type_id, ":ts": now}
            )
            logger.warning("Conflict on trusted mapping %s: existing_type=%s, new_type=%s",
                           mapping_key, existing.get("item_type_id"), item_type_id)
            return {"conflict": True, "existing_item_type_id": existing.get("item_type_id"),
                    "existing_category": existing.get("category")}

    # Overwrite or create — manual correction is immediately trusted
    table.put_item(Item={
        "mapping_key":     mapping_key,
        "store_id":        effective_store,
        "normalized_name": normalized_name,
        "item_type_id":    item_type_id,
        "category":        category,        # kept for backwards compat
        "confidence":      "1.00",
        "match_count":     1,
        "trust":           TrustLevel.TRUSTED.value,
        "source":          "manual",
        "created_at":      existing.get("created_at", now) if existing else now,
        "last_seen":       now
    })
    return {"conflict": False}


def load_store_mappings(store_id: str) -> list[dict]:
    """
    Load ALL mappings for a store via the StoreIndex GSI.
    Called at Lambda cold start to populate the in-memory fuzzy cache.
    Loads both store-scoped AND global mappings.
    """
    table  = db().Table(MAPPINGS_TABLE)
    result = []

    for sid in [store_id, GLOBAL_STORE]:
        if not sid or sid == "unknown":
            continue
        kwargs = {
            "IndexName": "StoreIndex",
            "KeyConditionExpression": Key("store_id").eq(sid)
        }
        while True:
            resp = table.query(**kwargs)
            result.extend(resp.get("Items", []))
            if not resp.get("LastEvaluatedKey"):
                break
            kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]

    logger.info("Loaded %d mappings for store=%s at cold start", len(result), store_id)
    return result


def list_all_mappings(store_id: str = None) -> list[dict]:
    """Return all mappings, optionally filtered by store_id."""
    table  = db().Table(MAPPINGS_TABLE)
    result = []
    if store_id:
        kwargs = {
            "IndexName": "StoreIndex",
            "KeyConditionExpression": Key("store_id").eq(store_id)
        }
        while True:
            resp = table.query(**kwargs)
            result.extend(resp.get("Items", []))
            if not resp.get("LastEvaluatedKey"):
                break
            kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    else:
        kwargs2: dict = {}
        while True:
            resp = table.scan(**kwargs2)
            result.extend(resp.get("Items", []))
            if not resp.get("LastEvaluatedKey"):
                break
            kwargs2["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    return result


def load_item_types() -> list[dict]:
    """Scan ItemTypes table for two-hop category resolution at cold start."""
    if not ITEM_TYPES_TABLE:
        return []
    table  = db().Table(ITEM_TYPES_TABLE)
    result = []
    kwargs: dict = {}
    while True:
        resp = table.scan(**kwargs)
        result.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    logger.info("Loaded %d item types", len(result))
    return result


def get_item_type(item_type_id: str) -> Optional[dict]:
    """Get a single item type by ID."""
    if not ITEM_TYPES_TABLE or not item_type_id:
        return None
    resp = db().Table(ITEM_TYPES_TABLE).get_item(Key={"item_type_id": item_type_id})
    return resp.get("Item")


def save_item_type(item_type_id: str, category_id: str, label: str = None) -> dict:
    """Create or update an item type entry."""
    table = db().Table(ITEM_TYPES_TABLE)
    now   = _now()
    existing = get_item_type(item_type_id)
    item = {
        "item_type_id": item_type_id,
        "category_id":  category_id,
        "label":        label or item_type_id,
        "created_at":   existing.get("created_at", now) if existing else now,
        "updated_at":   now
    }
    table.put_item(Item=item)
    logger.info("Saved item type %s -> %s", item_type_id, category_id)
    return item


def get_image_presigned_url(s3_key: str, expiry: int = 3600) -> str:
    """Generate a presigned GET URL for a receipt image in S3."""
    s3 = boto3.client("s3")
    return s3.generate_presigned_url(
        "get_object",
        Params={"Bucket": IMAGES_BUCKET, "Key": s3_key},
        ExpiresIn=expiry
    )


def set_item_expense(receipt_id: str, item_seq: str, is_expense: bool):
    """Mark or unmark a line item as a personal expense."""
    db().Table(ITEMS_TABLE).update_item(
        Key={"receipt_id": receipt_id, "item_seq": item_seq},
        UpdateExpression="SET is_expense = :v, updated_at = :t",
        ExpressionAttributeValues={":v": is_expense, ":t": _now()}
    )


def list_expense_items() -> list[dict]:
    """Return all items marked as expenses, joined with their receipt metadata."""
    from boto3.dynamodb.conditions import Attr

    # Scan items table for expense items
    table  = db().Table(ITEMS_TABLE)
    result = []
    kwargs: dict = {"FilterExpression": Attr("is_expense").eq(True)}
    while True:
        resp = table.scan(**kwargs)
        result.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]

    if not result:
        return []

    # Batch-get unique receipts for date / retailer / s3_key
    receipt_ids = list({item["receipt_id"] for item in result})
    receipts: dict = {}
    for i in range(0, len(receipt_ids), 100):
        chunk = receipt_ids[i:i + 100]
        resp = db().batch_get_item(RequestItems={
            RECEIPTS_TABLE: {"Keys": [{"receipt_id": rid} for rid in chunk]}
        })
        for r in resp.get("Responses", {}).get(RECEIPTS_TABLE, []):
            receipts[r["receipt_id"]] = r

    # Join receipt metadata onto each item
    enriched = []
    s3 = boto3.client("s3")
    for item in result:
        receipt  = receipts.get(item["receipt_id"], {})
        s3_key   = receipt.get("s3_key", "")
        image_url = None
        if s3_key and IMAGES_BUCKET:
            try:
                image_url = s3.generate_presigned_url(
                    "get_object",
                    Params={"Bucket": IMAGES_BUCKET, "Key": s3_key},
                    ExpiresIn=3600
                )
            except Exception:
                pass
        enriched.append({
            **item,
            "receipt_date":   receipt.get("receipt_date", ""),
            "retailer_name":  receipt.get("retailer_name") or receipt.get("retailer_id", ""),
            "image_url":      image_url,
        })

    enriched.sort(key=lambda x: x.get("receipt_date", ""), reverse=True)
    return enriched


# =============================================================================
# EXCEL BACKUP + UPDATE
#
# Order of operations (agreed v4):
#   1. Read master.xlsx from S3
#   2. Create dated backup FIRST: master_backup_YYYY-MM-DD.xlsx
#   3. Append new rows to in-memory workbook
#   4. Write updated master.xlsx back to S3
#   If step 4 fails, the backup from step 2 is already safe.
# =============================================================================

def backup_and_append_excel(exports_bucket: str, items: list[dict],
                             receipt_meta: dict) -> bool:
    """
    Backup master.xlsx then append new rows.
    Returns True on success, False on failure (backup preserved either way).
    Called by processor Lambda after saving items to DynamoDB.

    items: list of item dicts from grocery-items table
    receipt_meta: dict with receipt_date, retailer_name, total_amount
    """
    import io
    import openpyxl

    s3     = boto3.client("s3")
    today  = datetime.now(timezone.utc).date().isoformat()
    master = "master.xlsx"
    backup = f"master_backup_{today}.xlsx"

    # Step 1: Read existing master.xlsx (or create fresh if first run)
    try:
        obj = s3.get_object(Bucket=exports_bucket, Key=master)
        wb  = openpyxl.load_workbook(io.BytesIO(obj["Body"].read()))
        ws  = wb.active
    except s3.exceptions.NoSuchKey:
        # First ever run — create fresh workbook with headers
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Receipts"
        ws.append(["Date", "Store", "Item (Raw)", "Item (Normalised)",
                   "Category", "Price", "Confidence", "Match Source", "Receipt ID"])

    # Step 2: Create dated backup BEFORE making any changes
    backup_buffer = io.BytesIO()
    wb.save(backup_buffer)
    backup_buffer.seek(0)
    try:
        s3.put_object(Bucket=exports_bucket, Key=backup,
                      Body=backup_buffer.read(),
                      ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        logger.info("Backup created: %s", backup)
    except Exception as e:
        logger.error("Failed to create backup %s: %s", backup, e)
        return False  # abort — don't write master if backup failed

    # Step 3: Append new rows (one row per line item)
    receipt_date    = receipt_meta.get("receipt_date", today)
    retailer_name   = receipt_meta.get("retailer_name", "Unknown")
    receipt_id      = receipt_meta.get("receipt_id", "")

    for item in items:
        ws.append([
            receipt_date,
            retailer_name,
            item.get("raw_name", ""),
            item.get("normalized_name", ""),
            item.get("category", ""),
            item.get("price", ""),
            item.get("confidence", ""),
            item.get("match_source", ""),
            receipt_id
        ])

    # Step 4: Write updated master.xlsx back to S3
    master_buffer = io.BytesIO()
    wb.save(master_buffer)
    master_buffer.seek(0)
    try:
        s3.put_object(Bucket=exports_bucket, Key=master,
                      Body=master_buffer.read(),
                      ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        logger.info("master.xlsx updated with %d new rows", len(items))
        return True
    except Exception as e:
        logger.error("Failed to write updated master.xlsx: %s", e)
        return False  # backup from step 2 is still safe


def list_backups(exports_bucket: str) -> list[dict]:
    """List all backup files available for restore, newest first."""
    s3   = boto3.client("s3")
    resp = s3.list_objects_v2(Bucket=exports_bucket, Prefix="master_backup_")
    files = resp.get("Contents", [])
    return sorted(
        [{"filename": f["Key"], "size": f["Size"],
          "last_modified": f["LastModified"].isoformat()} for f in files],
        key=lambda x: x["last_modified"],
        reverse=True
    )


def restore_backup(exports_bucket: str, backup_filename: str) -> bool:
    """Copy a backup file back as the current master.xlsx."""
    s3 = boto3.client("s3")
    try:
        s3.copy_object(
            Bucket=exports_bucket,
            CopySource={"Bucket": exports_bucket, "Key": backup_filename},
            Key="master.xlsx"
        )
        logger.info("Restored %s → master.xlsx", backup_filename)
        return True
    except Exception as e:
        logger.error("Failed to restore backup %s: %s", backup_filename, e)
        return False


# =============================================================================
# REFERENCE DATA
# =============================================================================

def get_all_categories() -> list[dict]:
    resp = db().Table(CATEGORIES_TABLE).scan()
    return sorted(resp.get("Items", []), key=lambda x: x.get("name", ""))


def get_all_retailers() -> list[dict]:
    resp = db().Table(RETAILERS_TABLE).scan()
    return sorted(resp.get("Items", []), key=lambda x: x.get("name", ""))


def get_retailer(retailer_id: str) -> Optional[dict]:
    resp = db().Table(RETAILERS_TABLE).get_item(Key={"retailer_id": retailer_id})
    return resp.get("Item")


def detect_retailer(header_text: str) -> tuple[str, bool]:
    """
    Match receipt header text against known retailer header_patterns.
    Returns (retailer_id, identified). Falls back to ("unknown", False).
    """
    if not header_text:
        return "unknown", False
    text_lower = header_text.lower()
    retailers  = get_all_retailers()
    for r in retailers:
        for pattern in r.get("header_patterns", []):
            if pattern.lower() in text_lower:
                logger.info("Detected retailer: %s (pattern: %s)", r["retailer_id"], pattern)
                return r["retailer_id"], True
    logger.info("Could not detect retailer from header text")
    return "unknown", False


def save_retailer(retailer: dict):
    db().Table(RETAILERS_TABLE).put_item(Item=retailer)


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------
import io as _io
import re as _re_exp

MASTER_XLSX = "exports/master.xlsx"


def _parse_price(price_str):
    if not price_str:
        return 0.0
    cleaned = _re_exp.sub(r"[^\d.]", "", str(price_str))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _normalize_quantity(qty):
    """Normalise quantity strings: X4 -> 4, 0.500 kg -> 0.5, 500g -> 500."""
    import re
    if not qty:
        return qty
    qty = str(qty).strip()
    # X4, x4, 4X, 4x
    m = re.match(r'^[xX](\d+\.?\d*)$', qty) or re.match(r'^(\d+\.?\d*)[xX]$', qty)
    if m:
        val = float(m.group(1))
        return int(val) if val == int(val) else val
    # Number with unit suffix (0.500 kg, 500g, 1.5l, etc.)
    m = re.match(r'^(\d+\.?\d*)\s*(kg|g|ml|l|oz|lb)$', qty, re.IGNORECASE)
    if m:
        val = float(m.group(1))
        return int(val) if val == int(val) else val
    # Plain number
    try:
        val = float(qty)
        return int(val) if val == int(val) else val
    except ValueError:
        return qty


def receipt_export_ready(receipt_id: str) -> bool:
    items = get_items(receipt_id)
    if not items:
        return False
    for item in items:
        needs_review = item.get("needs_review", False)
        confirmed    = item.get("confirmed", False)
        if needs_review and not confirmed:
            return False
    return True


def export_receipt_to_excel(receipt_id: str, exports_bucket: str):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    receipt = get_receipt(receipt_id)
    items   = get_items(receipt_id)
    if not receipt or not items:
        return None

    s3c = boto3.client("s3")

    HEADERS = [
        "Date", "Store",
        "Item Name", "Item Type", "Category",
        "Price (GBP)", "Quantity",
        "Exported At", "Receipt ID"
    ]
    RID_COL = HEADERS.index("Receipt ID") + 1  # 1-based column index

    try:
        obj = s3c.get_object(Bucket=exports_bucket, Key=MASTER_XLSX)
        wb  = load_workbook(_io.BytesIO(obj["Body"].read()))
        ws  = wb.active
        existing_header = [c.value for c in ws[1]]
        if existing_header != HEADERS:
            # Legacy or mismatched format — wipe all data and rewrite header
            ws.delete_rows(1, ws.max_row)
            is_new = True
        else:
            is_new = False
            # Remove any existing rows for this receipt to prevent duplicates
            rows_to_delete = [
                row[0].row for row in ws.iter_rows(min_row=2)
                if ws.cell(row=row[0].row, column=RID_COL).value == receipt_id
            ]
            for row_num in reversed(rows_to_delete):
                ws.delete_rows(row_num)
    except Exception:
        wb     = Workbook()
        ws     = wb.active
        ws.title = "Grocery Items"
        is_new = True

    if is_new:
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2E4057")
            cell.alignment = Alignment(horizontal="center")
        widths = [12, 18, 35, 20, 18, 12, 10, 18, 36]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    raw_date   = receipt.get("receipt_date", "")
    store_name = receipt.get("retailer_name") or receipt.get("retailer_id", "")
    exported_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")

    # Parse receipt date to a proper date object for consistent Excel formatting
    receipt_date = raw_date
    try:
        from dateutil import parser as _dp
        s = str(raw_date)
        dayfirst = not s[:4].isdigit() or s[4:5] not in ("-", "/")
        receipt_date = _dp.parse(s, dayfirst=dayfirst).date()
    except Exception:
        pass

    exportable = [i for i in items if not i.get("is_expense")]
    for item in exportable:
        ws.append([
            receipt_date,
            store_name,
            item.get("raw_name", ""),
            item.get("item_type_id", ""),
            item.get("category", ""),
            _parse_price(item.get("price", "0")),
            _normalize_quantity(item.get("quantity", "1")),
            exported_at,
            receipt_id,
        ])

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3c.put_object(
        Bucket=exports_bucket,
        Key=MASTER_XLSX,
        Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    now = datetime.now(timezone.utc).isoformat()
    for item in exportable:
        db().Table(ITEMS_TABLE).update_item(
            Key={"receipt_id": receipt_id, "item_seq": item["item_seq"]},
            UpdateExpression="SET exported_to_excel = :t, exported_at = :ts, export_filename = :fn",
            ExpressionAttributeValues={":t": True, ":ts": now, ":fn": MASTER_XLSX}
        )

    # Mark receipt-level exported flag (for list view badge)
    update_receipt(receipt_id, exported_to_excel=True)

    logger.info(f"Exported {len(exportable)} items (skipped {len(items)-len(exportable)} expenses) from {receipt_id} to {MASTER_XLSX}")
    return MASTER_XLSX


def remove_receipt_from_excel(receipt_id: str, exports_bucket: str):
    """Remove all rows for a receipt from master.xlsx. No-op if receipt not found or no Receipt ID column."""
    from openpyxl import load_workbook
    s3c = boto3.client("s3")
    try:
        obj = s3c.get_object(Bucket=exports_bucket, Key=MASTER_XLSX)
        wb  = load_workbook(_io.BytesIO(obj["Body"].read()))
        ws  = wb.active
    except Exception:
        return  # no file yet — nothing to remove

    header = [c.value for c in ws[1]]
    if "Receipt ID" not in header:
        logger.warning("master.xlsx has no 'Receipt ID' column — skipping Excel cleanup for %s", receipt_id)
        return

    rid_col = header.index("Receipt ID") + 1  # 1-based
    rows_to_delete = [
        row[0].row for row in ws.iter_rows(min_row=2)
        if ws.cell(row=row[0].row, column=rid_col).value == receipt_id
    ]
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)

    if not rows_to_delete:
        logger.info("No rows found in master.xlsx for receipt %s", receipt_id)
        return

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3c.put_object(
        Bucket=exports_bucket,
        Key=MASTER_XLSX,
        Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Clear receipt-level exported flag
    try:
        update_receipt(receipt_id, exported_to_excel=False)
    except Exception:
        pass
    logger.info("Removed %d row(s) for receipt %s from master.xlsx", len(rows_to_delete), receipt_id)


# =============================================================================
# STARLING RECONCILIATION
# =============================================================================

def get_receipts_for_month(year_month: str) -> list[dict]:
    """Return all receipts whose receipt_date falls in YYYY-MM.
    Scans all receipts and filters in Python to handle varied date formats."""
    from dateutil import parser as _dp
    year, mo = map(int, year_month.split("-"))
    table    = db().Table(RECEIPTS_TABLE)
    result   = []
    kwargs: dict = {}
    while True:
        resp = table.scan(**kwargs)
        result.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    filtered = []
    for item in result:
        raw = item.get("receipt_date", "")
        if not raw or str(raw) == "None":
            continue
        try:
            s = str(raw)
            # ISO format (YYYY-MM-DD…) must NOT use dayfirst — it would swap month/day
            dayfirst = not s[:4].isdigit() or s[4:5] not in ("-", "/")
            d = _dp.parse(s, dayfirst=dayfirst)
            if d.year == year and d.month == mo:
                filtered.append(_decimals_to_float(item))
        except Exception:
            continue
    return filtered


def save_starling_transactions(month: str, transactions: list[dict]):
    """Overwrite all Starling transactions for a month in StarlingTable."""
    if not STARLING_TABLE:
        return
    table = db().Table(STARLING_TABLE)
    # Delete existing rows for this month first
    existing = get_starling_transactions(month)
    with table.batch_writer() as batch:
        for txn in existing:
            batch.delete_item(Key={"month": month, "transaction_id": txn["transaction_id"]})
    # Write new rows
    with table.batch_writer() as batch:
        for txn in transactions:
            batch.put_item(Item=_floats_to_decimal({**txn, "month": month}))
    logger.info("Saved %d Starling transactions for %s", len(transactions), month)


def get_starling_transactions(month: str) -> list[dict]:
    """Return all cached Starling transactions for a month, newest first."""
    if not STARLING_TABLE:
        return []
    table  = db().Table(STARLING_TABLE)
    kwargs = {"KeyConditionExpression": Key("month").eq(month)}
    items  = []
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    return sorted(
        [_decimals_to_float(i) for i in items],
        key=lambda x: x.get("date", ""),
        reverse=True,
    )


def update_starling_transaction_match(
    month: str, transaction_id: str, receipt_id, match_status: str,
    receipt_amount=None, diff_amount=None
):
    """Update the match fields on a single Starling transaction (manual link/unlink)."""
    if not STARLING_TABLE:
        return
    table = db().Table(STARLING_TABLE)
    expr_names  = {
        "#ms": "match_status", "#mr": "matched_receipt_id",
        "#ra": "receipt_amount", "#da": "diff_amount", "#sa": "synced_at",
    }
    expr_values = {
        ":ms": match_status, ":mr": receipt_id,
        ":ra": receipt_amount, ":da": diff_amount, ":sa": _now(),
    }
    table.update_item(
        Key={"month": month, "transaction_id": transaction_id},
        UpdateExpression="SET #ms = :ms, #mr = :mr, #ra = :ra, #da = :da, #sa = :sa",
        ExpressionAttributeNames=expr_names,
        ExpressionAttributeValues=_floats_to_decimal(expr_values),
    )
