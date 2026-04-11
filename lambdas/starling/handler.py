"""
Starling Lambda — syncs Starling Bank transactions and reconciles with grocery receipts.

Routes:
  POST /starling/sync?month=YYYY-MM        — fetch, match, store, write Excel
  GET  /starling/transactions?month=YYYY-MM — read cached results from DynamoDB
"""

import io
import json
import os
import re
import urllib.request
import urllib.error
from calendar import monthrange
from datetime import datetime, timezone, date

import boto3
import openpyxl
from dateutil import parser as dateutil_parser
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import dynamo_client as db

EXPORTS_BUCKET    = os.environ["EXPORTS_BUCKET"]
ENVIRONMENT       = os.environ.get("ENVIRONMENT", "dev")
_ORIGIN           = os.environ.get("FRONTEND_ORIGIN", "*")

STARLING_API_BASE = "https://api.starlingbank.com"
STARLING_EXCEL    = "exports/starling-reconciliation.xlsx"

_ssm_cache: dict = {}


def _get_ssm(name: str) -> str:
    if name not in _ssm_cache:
        ssm = boto3.client("ssm", region_name="eu-west-2")
        _ssm_cache[name] = ssm.get_parameter(
            Name=name, WithDecryption=True
        )["Parameter"]["Value"]
    return _ssm_cache[name]


def _starling_get(path: str, token: str) -> dict:
    url = f"{STARLING_API_BASE}{path}"
    req = urllib.request.Request(
        url,
        headers={"Authorization": f"Bearer {token}", "Accept": "application/json"}
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def lambda_handler(event, context):
    method = event.get("httpMethod", "")
    path   = event.get("path", "")
    qp     = event.get("queryStringParameters") or {}

    try:
        body = json.loads(event.get("body") or "{}")
        if method == "POST" and "sync" in path:
            return _sync(qp)
        elif method == "GET" and "transactions" in path:
            return _get_transactions(qp)
        elif method == "GET" and "export" in path:
            return _get_export_url()
        elif method == "GET" and "receipts" in path:
            return _get_month_receipts(qp)
        elif method == "POST" and "match" in path:
            return _manual_match(body)
        elif method == "DELETE" and "match" in path:
            return _unlink_match(body)
        return _error(404, "Not found")
    except Exception as e:
        import traceback; traceback.print_exc()
        return _error(500, str(e))


def _sync(qp: dict) -> dict:
    """Fetch Starling transactions, match against receipts, store in DDB, write Excel."""
    month = qp.get("month") or datetime.now(timezone.utc).strftime("%Y-%m")
    try:
        year, mo = map(int, month.split("-"))
    except ValueError:
        return _error(400, "Invalid month — use YYYY-MM")

    _, last_day = monthrange(year, mo)
    min_ts = f"{month}-01T00:00:00.000Z"
    max_ts = f"{month}-{last_day:02d}T23:59:59.999Z"

    token  = _get_ssm(f"/grocery-scanner/{ENVIRONMENT}/starling-token")
    spaces = json.loads(_get_ssm(f"/grocery-scanner/{ENVIRONMENT}/starling-spaces"))

    # Get primary account UID
    accounts = _starling_get("/api/v2/accounts", token).get("accounts", [])
    primary  = next((a for a in accounts if a.get("accountType") == "PRIMARY"), None)
    if not primary:
        return _error(500, "No primary Starling account found")
    account_uid = primary["accountUid"]

    # Load receipts for this month — deduplicate by (retailer, amount, date)
    receipts = db.get_receipts_for_month(month)
    seen: set = set()
    deduped: list = []
    for r in receipts:
        key = (
            _normalise_name(r.get("retailer_name") or ""),
            str(r.get("total_amount", "")),
            str(r.get("receipt_date", "")),
        )
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    available_receipts = deduped

    # Collect all outgoing transactions across all spaces first
    pending = []
    for space in spaces:
        space_name = space["name"]
        space_uid  = space["uid"]
        path = (
            f"/api/v2/feed/account/{account_uid}/category/{space_uid}"
            f"/transactions-between"
            f"?minTransactionTimestamp={min_ts}&maxTransactionTimestamp={max_ts}"
        )
        txns = _starling_get(path, token).get("feedItems", [])
        for t in txns:
            if t.get("direction") == "OUT":
                pending.append({
                    "txn":        t,
                    "space_name": space_name,
                    "txn_date":   dateutil_parser.parse(t["transactionTime"]).date(),
                    "txn_amount": t["amount"]["minorUnits"] / 100,
                    "merchant":   t.get("counterPartyName", ""),
                })

    # Pass 1: exact matches (amount ±1p + date ±2d) — globally before any partial
    exact: dict[int, tuple] = {}
    for i, p in enumerate(pending):
        r = _exact_match(p["txn_date"], p["txn_amount"], available_receipts)
        if r:
            available_receipts.remove(r)
            exact[i] = r

    # Pass 2: partial matches (name overlap + date ±2d) for remaining unmatched
    partial: dict[int, tuple] = {}
    for i, p in enumerate(pending):
        if i in exact:
            continue
        r = _partial_match(p["txn_date"], p["merchant"], available_receipts)
        if r:
            available_receipts.remove(r)
            partial[i] = r

    # Load existing manual matches so re-sync doesn't overwrite them
    existing_txns = {t["transaction_id"]: t for t in db.get_starling_transactions(month)}

    # Build result rows
    all_rows = []
    synced_at = datetime.now(timezone.utc).isoformat()
    for i, p in enumerate(pending):
        if i in exact:
            matched_r, status = exact[i], "matched"
        elif i in partial:
            matched_r, status = partial[i], "partial"
        else:
            matched_r, status = None, "unmatched"
        r_amount = _parse_amount(matched_r.get("total_amount")) if matched_r else None
        diff     = round(p["txn_amount"] - r_amount, 2) if r_amount is not None else None
        txn_id   = p["txn"]["feedItemUid"]

        # Preserve any manual link the user made in a previous sync
        prev = existing_txns.get(txn_id, {})
        if prev.get("match_status") == "manual":
            status    = "manual"
            matched_r = None  # we don't re-look up the receipt object; use stored values
            r_amount  = prev.get("receipt_amount")
            diff      = prev.get("diff_amount")
            manual_receipt_id = prev.get("matched_receipt_id")
        else:
            manual_receipt_id = None

        all_rows.append({
            "month":              month,
            "transaction_id":     txn_id,
            "date":               p["txn_date"].isoformat(),
            "merchant_name":      p["merchant"],
            "space_name":         p["space_name"],
            "amount":             p["txn_amount"],
            "match_status":       status,
            "matched_receipt_id": manual_receipt_id if status == "manual" else (matched_r.get("receipt_id") if matched_r else None),
            "receipt_amount":     r_amount,
            "diff_amount":        diff,
            "synced_at":          synced_at,
        })

    db.save_starling_transactions(month, all_rows)
    _write_starling_excel(month, all_rows)

    matched   = sum(1 for r in all_rows if r["match_status"] in ("matched", "manual"))
    partial   = sum(1 for r in all_rows if r["match_status"] == "partial")
    unmatched = sum(1 for r in all_rows if r["match_status"] == "unmatched")

    return _ok({
        "month":     month,
        "total":     len(all_rows),
        "matched":   matched,
        "partial":   partial,
        "unmatched": unmatched,
        "synced_at": datetime.now(timezone.utc).isoformat(),
    })


def _get_export_url() -> dict:
    """Return a presigned S3 URL to download starling-reconciliation.xlsx."""
    s3 = boto3.client("s3")
    try:
        url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": EXPORTS_BUCKET, "Key": STARLING_EXCEL},
            ExpiresIn=300,
        )
        return _ok({"url": url})
    except Exception:
        return _error(404, "No export file found — run a sync first")


def _get_month_receipts(qp: dict) -> dict:
    """Return unmatched receipts for a month — used by the manual-match receipt picker.
    Excludes receipts already linked to a Starling transaction this month."""
    month    = qp.get("month") or datetime.now(timezone.utc).strftime("%Y-%m")
    txn_id   = qp.get("exclude_txn")  # optional: the transaction being re-linked (keep its receipt available)
    receipts = db.get_receipts_for_month(month)

    # Collect receipt IDs already matched to other transactions this month
    txns = db.get_starling_transactions(month)
    matched_ids = {
        t["matched_receipt_id"]
        for t in txns
        if t.get("matched_receipt_id") and t.get("transaction_id") != txn_id
    }

    rows = [
        {
            "receipt_id":    r.get("receipt_id"),
            "receipt_date":  r.get("receipt_date"),
            "retailer_name": r.get("retailer_name") or r.get("retailer_id", ""),
            "total_amount":  r.get("total_amount"),
        }
        for r in receipts
        if r.get("receipt_id") not in matched_ids
    ]
    rows.sort(key=lambda r: str(r.get("receipt_date") or ""), reverse=True)
    return _ok({"month": month, "receipts": rows})


def _manual_match(body: dict) -> dict:
    """Manually link a Starling transaction to a receipt."""
    month          = body.get("month")
    transaction_id = body.get("transaction_id")
    receipt_id     = body.get("receipt_id")
    if not all([month, transaction_id, receipt_id]):
        return _error(400, "month, transaction_id and receipt_id are required")
    # Look up the receipt total so receipt_amount is stored correctly
    receipt   = db.get_receipt(receipt_id) or {}
    r_amount  = _parse_amount(receipt.get("total_amount"))
    txn_amount = None
    for t in db.get_starling_transactions(month):
        if t.get("transaction_id") == transaction_id:
            txn_amount = t.get("amount")
            break
    diff = round(txn_amount - r_amount, 2) if (txn_amount is not None and r_amount is not None) else None
    db.update_starling_transaction_match(month, transaction_id, receipt_id, "manual", r_amount, diff)
    _regenerate_excel(month)
    return _ok({"updated": True})


def _unlink_match(body: dict) -> dict:
    """Remove a manual (or any) match from a Starling transaction."""
    month          = body.get("month")
    transaction_id = body.get("transaction_id")
    if not all([month, transaction_id]):
        return _error(400, "month and transaction_id are required")
    db.update_starling_transaction_match(month, transaction_id, None, "unmatched")
    _regenerate_excel(month)
    return _ok({"updated": True})


def _regenerate_excel(month: str):
    """Re-write the Excel sheet for a month using current DynamoDB state."""
    rows = db.get_starling_transactions(month)
    if rows:
        _write_starling_excel(month, rows)


def _get_transactions(qp: dict) -> dict:
    """Return cached Starling transactions from DynamoDB."""
    month = qp.get("month") or datetime.now(timezone.utc).strftime("%Y-%m")
    rows  = db.get_starling_transactions(month)

    # Enrich matched rows with items_total (sum of non-expense item prices)
    for row in rows:
        rid = row.get("matched_receipt_id")
        if rid:
            items = db.get_items(rid)
            total = sum(
                _parse_amount(i.get("price")) or 0
                for i in items
                if not i.get("is_expense")
            )
            row["items_total"] = round(total, 2) if items else None
        else:
            row["items_total"] = None

    matched   = sum(1 for r in rows if r.get("match_status") in ("matched", "manual"))
    partial   = sum(1 for r in rows if r.get("match_status") == "partial")
    unmatched = sum(1 for r in rows if r.get("match_status") == "unmatched")
    return _ok({
        "month":        month,
        "transactions": rows,
        "total":        len(rows),
        "matched":      matched,
        "partial":      partial,
        "unmatched":    unmatched,
    })


def _normalise_name(s: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", "", s.lower())).strip()


def _names_overlap(merchant: str, retailer: str) -> bool:
    if not merchant or not retailer:
        return False
    m = _normalise_name(merchant)
    r = _normalise_name(retailer)
    m_words = [w for w in m.split() if len(w) > 2]
    r_words = [w for w in r.split() if len(w) > 2]
    if not m_words or not r_words:
        return False
    return any(w in r for w in m_words) or any(w in m for w in r_words)


def _exact_match(txn_date: date, txn_amount: float, receipts: list):
    """Return receipt with exact amount (±1p) and date ±10d, or None."""
    for r in receipts:
        r_date   = _parse_date(r.get("receipt_date"))
        r_amount = _parse_amount(r.get("total_amount"))
        if (r_date and r_amount is not None
                and abs((txn_date - r_date).days) <= 10
                and abs(txn_amount - r_amount) < 0.01):
            return r
    return None


def _partial_match(txn_date: date, merchant: str, receipts: list):
    """Return receipt with matching merchant name, date ±2d, and a known total, or None.
    Receipts with no total_amount are excluded — they can't be meaningfully reconciled."""
    for r in receipts:
        r_date   = _parse_date(r.get("receipt_date"))
        retailer = r.get("retailer_name", "") or ""
        if (r_date
                and _parse_amount(r.get("total_amount")) is not None
                and abs((txn_date - r_date).days) <= 2
                and _names_overlap(merchant, retailer)):
            return r
    return None


def _parse_amount(val) -> float | None:
    if val is None:
        return None
    s = re.sub(r"[^\d.]", "", str(val))
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(val) -> date | None:
    if not val or str(val) == "None":
        return None
    try:
        s = str(val)
        # ISO format (YYYY-MM-DD…) must NOT use dayfirst — it would swap month/day
        dayfirst = not s[:4].isdigit() or s[4:5] not in ("-", "/")
        return dateutil_parser.parse(s, dayfirst=dayfirst).date()
    except Exception:
        return None


def _write_starling_excel(month: str, rows: list):
    """Write (overwrite) the month sheet in starling-reconciliation.xlsx on S3."""
    s3c = boto3.client("s3")

    try:
        obj = s3c.get_object(Bucket=EXPORTS_BUCKET, Key=STARLING_EXCEL)
        wb  = openpyxl.load_workbook(io.BytesIO(obj["Body"].read()))
    except Exception:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if month in wb.sheetnames:
        del wb[month]
    ws = wb.create_sheet(title=month)

    GREEN  = PatternFill("solid", fgColor="D6EAD6")
    YELLOW = PatternFill("solid", fgColor="FFF3CD")
    RED    = PatternFill("solid", fgColor="FADADD")
    BLUE   = PatternFill("solid", fgColor="D6EAF8")
    HEADER = PatternFill("solid", fgColor="2A6049")
    thin   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    headers = ["Date", "Shop", "Space", "Starling GBP", "Receipt GBP", "Diff GBP", "Status", "Receipt ID"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill      = HEADER
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin

    for row in sorted(rows, key=lambda r: r["date"], reverse=True):
        status = row["match_status"]
        fill   = GREEN if status == "matched" else YELLOW if status == "partial" else BLUE if status == "manual" else RED
        ws.append([
            row["date"],
            row["merchant_name"],
            row["space_name"],
            row["amount"],
            row["receipt_amount"],
            row["diff_amount"],
            status,
            row.get("matched_receipt_id") or "",
        ])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(ws.max_row, col)
            cell.fill   = fill
            cell.border = thin
            if col in (4, 5, 6):
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")

    for i, w in enumerate([12, 28, 18, 14, 14, 12, 12, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3c.put_object(
        Bucket=EXPORTS_BUCKET,
        Key=STARLING_EXCEL,
        Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _ok(data: dict) -> dict:
    return {
        "statusCode": 200,
        "headers": {
            "Content-Type": "application/json",
            "Access-Control-Allow-Origin": _ORIGIN,
            "Access-Control-Allow-Headers": "Content-Type,Authorization",
        },
        "body": json.dumps({"success": True, **data}, default=str),
    }


def _error(status: int, message: str) -> dict:
    return {
        "statusCode": status,
        "headers": {
            "Content-Type": "application/json",
            "Access-Control-Allow-Origin": _ORIGIN,
        },
        "body": json.dumps({"success": False, "error": message}),
    }
