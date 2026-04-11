"""
SheetSync Lambda — triggered by S3 ObjectCreated on exports/master.xlsx.
Full-replace sync: reads Excel from S3, clears Google Sheet, writes all rows.
"""

import io
import json
import logging
import os

import boto3
from openpyxl import load_workbook
from google.oauth2 import service_account
from googleapiclient.discovery import build

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

SHEET_ID  = os.environ["GOOGLE_SHEET_ID"]
CREDS_SSM = os.environ["GOOGLE_CREDENTIALS_SSM"]
SCOPES    = ["https://www.googleapis.com/auth/spreadsheets"]

_creds_json = None   # cached service account dict (survives warm starts)
_sheets_svc = None   # cached Sheets API client


def _get_sheets_service():
    global _creds_json, _sheets_svc
    if _sheets_svc is None:
        if _creds_json is None:
            ssm  = boto3.client("ssm")
            resp = ssm.get_parameter(Name=CREDS_SSM, WithDecryption=False)
            _creds_json = json.loads(resp["Parameter"]["Value"])
        creds = service_account.Credentials.from_service_account_info(
            _creds_json, scopes=SCOPES
        )
        _sheets_svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
    return _sheets_svc


def _read_excel(bucket: str, key: str) -> list:
    s3  = boto3.client("s3")
    obj = s3.get_object(Bucket=bucket, Key=key)
    wb  = load_workbook(io.BytesIO(obj["Body"].read()), data_only=True)
    ws  = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        converted = []
        for cell in row:
            if cell is None:
                converted.append("")
            elif hasattr(cell, "isoformat"):   # date / datetime object
                converted.append(cell.isoformat())
            elif isinstance(cell, (int, float)):
                converted.append(cell)
            else:
                converted.append(str(cell))
        rows.append(converted)
    return rows


def _sync(bucket: str, key: str):
    rows = _read_excel(bucket, key)
    if not rows:
        logger.info("Empty Excel — nothing to sync")
        return

    svc = _get_sheets_service()
    sh  = svc.spreadsheets()

    # Full replace: clear all existing data then write everything fresh
    sh.values().clear(spreadsheetId=SHEET_ID, range="A:Z").execute()
    sh.values().update(
        spreadsheetId=SHEET_ID,
        range="A1",
        valueInputOption="USER_ENTERED",
        body={"values": rows},
    ).execute()

    logger.info("Synced %d data row(s) to Google Sheet %s", len(rows) - 1, SHEET_ID)


def lambda_handler(event, context):
    for record in event.get("Records", []):
        if record.get("eventSource") != "aws:s3":
            continue
        bucket = record["s3"]["bucket"]["name"]
        key    = record["s3"]["object"]["key"]
        logger.info("Syncing s3://%s/%s → Google Sheet", bucket, key)
        try:
            _sync(bucket, key)
        except Exception:
            import traceback
            traceback.print_exc()
            # Don't raise — prevents infinite S3 retry on transient Google API errors

    return {"statusCode": 200, "body": "ok"}
