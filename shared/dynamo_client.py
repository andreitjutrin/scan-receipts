"""
DynamoDB helpers shared by all Lambda functions.
Implements hybrid store-scoped + global fallback mapping lookups.

Confidence thresholds (v4 agreed):
  >= 0.92  silent, confident mapping
  0.75â€“0.91 flag review, tentative mapping
  0.50â€“0.74 flag review, no mapping written (too uncertain to learn from)
  < 0.50   unknown, no mapping written
"""

import os
import logging
from datetime import datetime, timezone
from typing import Optional

import boto3
from boto3.dynamodb.conditions import Key
from decimal import Decimal

from models import (
    THRESHOLD_SILENT, THRESHOLD_REVIEW, THRESHOLD_GUESS,
    TRUST_PROMOTE_CONFIDENT, TRUST_PROMOTE_TRUSTED, MAX_PROCESS_COUNT,
    TrustLevel
)

logger = logging.getLogger(__name__)

RECEIPTS_TABLE   = os.environ["RECEIPTS_TABLE"]
ITEMS_TABLE      = os.environ["ITEMS_TABLE"]
MAPPINGS_TABLE   = os.environ["MAPPINGS_TABLE"]
CATEGORIES_TABLE = os.environ["CATEGORIES_TABLE"]
RETAILERS_TABLE  = os.environ["RETAILERS_TABLE"]
ALERT_TOPIC_ARN  = os.environ.get("ALERT_TOPIC_ARN", "")  # SNS topic for cost alerts

GLOBAL_STORE = "global"

_db  = None
_sns = None


def db():
    global _db
    if _db is None:
        _db = boto3.resource("dynamodb")
    return _db


def sns():
    global _sns
    if _sns is None:
        _sns = boto3.client("sns")
    return _sns


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _make_key(store_id: str, normalized_name: str) -> str:
    return f"{store_id}#{normalized_name}"


# =============================================================================
# RECEIPTS
# =============================================================================

def save_receipt(receipt: dict):
    db().Table(RECEIPTS_TABLE).put_item(Item=receipt)


def get_receipt(receipt_id: str) -> Optional[dict]:
    resp = db().Table(RECEIPTS_TABLE).get_item(Key={"receipt_id": receipt_id})
    item = resp.get("Item")
    return _decimals_to_float(item) if item else None


def update_receipt(receipt_id: str, **fields):
    """Update arbitrary fields on a receipt record."""
    table = db().Table(RECEIPTS_TABLE)
    fields["updated_at"] = _now()

    set_parts  = [f"#f{i} = :v{i}" for i, k in enumerate(fields)]
    attr_names = {f"#f{i}": k for i, k in enumerate(fields)}
    attr_vals  = {f":v{i}": v for i, (k, v) in enumerate(fields.items())}

    table.update_item(
        Key={"receipt_id": receipt_id},
        UpdateExpression="SET " + ", ".join(set_parts),
        ExpressionAttributeNames=attr_names,
        ExpressionAttributeValues=attr_vals
    )


def list_receipts(user_id: str = "default",
                  date_from: str = None,
                  date_to:   str = None) -> list[dict]:
    table  = db().Table(RECEIPTS_TABLE)
    kwargs = {
        "IndexName": "UserDateIndex",
        "KeyConditionExpression": Key("user_id").eq(user_id),
        "ScanIndexForward": False
    }
    if date_from and date_to:
        kwargs["KeyConditionExpression"] &= Key("receipt_date").between(date_from, date_to)

    items = []
    while True:
        resp = table.query(**kwargs)
        items.extend(resp.get("Items", []))
        if not resp.get("LastEvaluatedKey"):
            break
        kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
    return [_decimals_to_float(i) for i in items]


def delete_receipt(receipt_id: str):
    db().Table(RECEIPTS_TABLE).delete_item(Key={"receipt_id": receipt_id})


def check_and_increment_process_count(receipt_id: str) -> bool:
    """
    Cost safeguard: atomically check and increment process_count.
    Returns True if processing is allowed (count was < MAX_PROCESS_COUNT).
    Returns False if limit reached â€” caller must stop and set status=failed.

    Uses a conditional update so this is safe even if two Lambda instances
    run simultaneously (which shouldn't happen with concurrency=2, but safe anyway).
    """
    table = db().Table(RECEIPTS_TABLE)
    try:
        table.update_item(
            Key={"receipt_id": receipt_id},
            UpdateExpression="SET process_count = if_not_exists(process_count, :zero) + :inc, "
                             "updated_at = :ts",
            ConditionExpression="attribute_not_exists(process_count) OR process_count < :max",
            ExpressionAttributeValues={
                ":zero": 0,
                ":inc":  1,
                ":max":  MAX_PROCESS_COUNT,
                ":ts":   _now()
            }
        )
        return True
    except db().meta.client.exceptions.ConditionalCheckFailedException:
        # process_count >= MAX_PROCESS_COUNT
        logger.error("COST SAFEGUARD: receipt %s hit MAX_PROCESS_COUNT=%d", receipt_id, MAX_PROCESS_COUNT)
        _send_alert(f"Receipt {receipt_id} hit max reprocess limit ({MAX_PROCESS_COUNT}). "
                    f"Processing stopped. Check CloudWatch logs for details.")
        return False


def _send_alert(message: str):
    """Send SNS alert email. Silently skips if topic not configured."""
    if not ALERT_TOPIC_ARN:
        logger.warning("ALERT (no SNS topic configured): %s", message)
        return
    try:
        sns().publish(
            TopicArn=ALERT_TOPIC_ARN,
            Subject="Grocery Scanner Alert",
            Message=message
        )
    except Exception as e:
        logger.error("Failed to send SNS alert: %s", e)


# =============================================================================
# ITEMS
# =============================================================================

def _floats_to_decimal(obj):
    """Recursively convert float values to Decimal for DynamoDB compatibility."""
    if isinstance(obj, float):
        return Decimal(str(obj))
    elif isinstance(obj, dict):
        return {k: _floats_to_decimal(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_floats_to_decimal(v) for v in obj]
    return obj


def save_items(receipt_id: str, items: list[dict]):
    table = db().Table(ITEMS_TABLE)
    with table.batch_writer() as batch:
        for item in items:
            clean = _floats_to_decimal(item)
            batch.put_item(Item={"receipt_id": receipt_id, **clean})


def _decimals_to_float(obj):
    """Recursively convert Decimal values back to float for JSON serialisation."""
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {k: _decimals_to_float(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_decimals_to_float(v) for v in obj]
    return obj


def get_items(receipt_id: str) -> list[dict]:
    resp = db().Table(ITEMS_TABLE).query(
        KeyConditionExpression=Key("receipt_id").eq(receipt_id)
    )
    items = resp.get("Items", [])
    items = [_decimals_to_float(i) for i in items]
    return sorted(items, key=lambda x: x.get("item_seq", ""))


def update_item_category(receipt_id: str, item_seq: str, category: str, record_correction: bool = False):
    expr = "SET category = :cat, confirmed = :yes, match_source = :src, updated_at = :ts"
    vals = {
        ":cat": category,
        ":yes": True,
        ":src": "manual",
        ":ts":  _now()
    }
    if record_correction:
        expr += ", corrected_at = :cor"
        vals[":cor"] = _now()
    db().Table(ITEMS_TABLE).update_item(
        Key={"receipt_id": receipt_id, "item_seq": item_seq},
        UpdateExpression=expr,
        ExpressionAttributeValues=vals
    )


def delete_items(receipt_id: str):
    """Delete all items for a receipt â€” used before reprocessing."""
    items = get_items(receipt_id)
    table = db().Table(ITEMS_TABLE)
    with table.batch_writer() as batch:
        for item in items:
            batch.delete_item(Key={
                "receipt_id": receipt_id,
                "item_seq":   item["item_seq"]
            })


# =============================================================================
# MAPPINGS â€” hybrid store-scoped + global fallback
#
# Key format:  "{store_id}#{normalized_name}"
#   "tesco#smoked salmon"   â†’ store-scoped
#   "global#smoked salmon"  â†’ global fallback (seeded)
#
# One BatchGetItem call checks both store-scoped and global simultaneously.
# Store-scoped result takes priority over global if both exist.
#
# After any fuzzy/partial match succeeds, the result is written back as a
# store-scoped mapping so future scans hit Layer 1 (exact) instead.
#
# Only mappings with confidence >= THRESHOLD_REVIEW (0.75) are persisted.
# Below that threshold the match is too uncertain to learn from.
# =============================================================================

def get_mapping_hybrid(store_id: str, normalized_name: str) -> Optional[dict]:
    """
    Single BatchGetItem that checks store-scoped AND global simultaneously.
    Returns the store-scoped mapping if found, else global, else None.
    """
    keys_to_fetch = [{"mapping_key": _make_key(GLOBAL_STORE, normalized_name)}]

    if store_id and store_id not in (GLOBAL_STORE, "unknown"):
        keys_to_fetch.insert(0, {"mapping_key": _make_key(store_id, normalized_name)})

    resp  = db().batch_get_item(RequestItems={MAPPINGS_TABLE: {"Keys": keys_to_fetch}})
    items = resp.get("Responses", {}).get(MAPPINGS_TABLE, [])

    store_hit  = next((i for i in items if i.get("store_id") == store_id),       None)
    global_hit = next((i for i in items if i.get("store_id") == GLOBAL_STORE),   None)

    return store_hit or global_hit


def promote_mapping_if_ready(mapping_key: str):
    """
    Increment match_count and promote trust level if thresholds are met.
    tentative â†’(3Ã—)â†’ confident â†’(5Ã—)â†’ trusted
    Trusted mappings are never demoted.
    """
    table = db().Table(MAPPINGS_TABLE)

    resp = table.update_item(
        Key={"mapping_key": mapping_key},
        UpdateExpression=(
            "SET match_count = if_not_exists(match_count, :zero) + :inc, "
            "last_seen = :ts"
        ),
        ExpressionAttributeValues={":zero": 0, ":inc": 1, ":ts": _now()},
        ReturnValues="ALL_NEW"
    )
    updated   = resp.get("Attributes", {})
    count     = int(updated.get("match_count", 0))
    trust     = updated.get("trust", "tentative")

    new_trust = None
    if trust == "tentative" and count >= TRUST_PROMOTE_CONFIDENT:
        new_trust = TrustLevel.CONFIDENT.value
    elif trust == "confident" and count >= TRUST_PROMOTE_TRUSTED:
        new_trust = TrustLevel.TRUSTED.value

    if new_trust:
        table.update_item(
            Key={"mapping_key": mapping_key},
            UpdateExpression="SET trust = :t",
            ExpressionAttributeValues={":t": new_trust}
        )
        logger.info("Promoted mapping %s â†’ %s", mapping_key, new_trust)


def write_learned_mapping(store_id: str, normalized_name: str,
                           category: str, confidence: float,
                           source: str) -> None:
    """
    Persist a fuzzy match result to the store-scoped table so future identical
    OCR strings hit Layer 1 (exact) rather than going through fuzzy matching again.

    Rules:
    - confidence >= 0.92 â†’ write as confident
    - confidence >= 0.75 â†’ write as tentative
    - confidence <  0.75 â†’ do NOT write (too uncertain to learn from)
    - never overwrite a trusted mapping with a fuzzy result
    - never write if store is unknown (don't know which template to update)
    """
    if confidence < THRESHOLD_REVIEW:
        return  # below 0.75 â€” don't learn from this
    if not store_id or store_id == "unknown":
        return

    trust       = TrustLevel.CONFIDENT.value if confidence >= THRESHOLD_SILENT else TrustLevel.TENTATIVE.value
    mapping_key = _make_key(store_id, normalized_name)
    now         = _now()

    existing = db().Table(MAPPINGS_TABLE).get_item(
        Key={"mapping_key": mapping_key}
    ).get("Item")

    if existing and existing.get("trust") == TrustLevel.TRUSTED.value:
        return  # never overwrite trusted with a fuzzy result

    db().Table(MAPPINGS_TABLE).put_item(Item={
        "mapping_key":     mapping_key,
        "store_id":        store_id,
        "normalized_name": normalized_name,
        "category":        category,
        "confidence":      str(round(confidence, 4)),
        "match_count":     1,
        "trust":           trust,
        "source":          source,
        "created_at":      existing.get("created_at", now) if existing else now,
        "last_seen":       now
    })
    logger.info("Learned mapping %s â†’ %s (%.0f%%, %s)", mapping_key, category, confidence * 100, trust)


def save_correction(store_id: str, normalized_name: str, category: str) -> dict:
    """
    Save a user correction back to the mappings table.

    Behaviour by existing trust level:
    - trusted + same category  â†’ just increment match_count, no change
    - trusted + different category â†’ save conflict flag, return conflict info
    - confident or tentative â†’ overwrite with manual correction (instantly trusted)
    - missing â†’ create new trusted mapping

    Returns dict with key 'conflict': True/False
    """
    effective_store = store_id if (store_id and store_id != "unknown") else GLOBAL_STORE
    mapping_key     = _make_key(effective_store, normalized_name)
    table           = db().Table(MAPPINGS_TABLE)
    now             = _now()

    existing = table.get_item(Key={"mapping_key": mapping_key}).get("Item")

    if existing and existing.get("trust") == TrustLevel.TRUSTED.value:
        if existing.get("category") == category:
            # User confirmed â€” just increment count
            promote_mapping_if_ready(mapping_key)
            return {"conflict": False}
        else:
            # Conflict â€” save it but don't overwrite
            table.update_item(
                Key={"mapping_key": mapping_key},
                UpdateExpression="SET conflict_category = :c, conflict_at = :ts",
                ExpressionAttributeValues={":c": category, ":ts": now}
            )
            logger.warning("Conflict on trusted mapping %s: existing=%s, correction=%s",
                           mapping_key, existing.get("category"), category)
            return {"conflict": True, "existing_category": existing.get("category")}

    # Overwrite or create â€” manual correction is immediately trusted
    table.put_item(Item={
        "mapping_key":     mapping_key,
        "store_id":        effective_store,
        "normalized_name": normalized_name,
        "category":        category,
        "confidence":      "1.00",
        "match_count":     1,
        "trust":           TrustLevel.TRUSTED.value,
        "source":          "manual",
        "created_at":      existing.get("created_at", now) if existing else now,
        "last_seen":       now
    })
    return {"conflict": False}


def load_store_mappings(store_id: str) -> list[dict]:
    """
    Load ALL mappings for a store via the StoreIndex GSI.
    Called at Lambda cold start to populate the in-memory fuzzy cache.
    Loads both store-scoped AND global mappings.
    """
    table  = db().Table(MAPPINGS_TABLE)
    result = []

    for sid in [store_id, GLOBAL_STORE]:
        if not sid or sid == "unknown":
            continue
        kwargs = {
            "IndexName": "StoreIndex",
            "KeyConditionExpression": Key("store_id").eq(sid)
        }
        while True:
            resp = table.query(**kwargs)
            result.extend(resp.get("Items", []))
            if not resp.get("LastEvaluatedKey"):
                break
            kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]

    logger.info("Loaded %d mappings for store=%s at cold start", len(result), store_id)
    return result


# =============================================================================
# EXCEL BACKUP + UPDATE
#
# Order of operations (agreed v4):
#   1. Read master.xlsx from S3
#   2. Create dated backup FIRST: master_backup_YYYY-MM-DD.xlsx
#   3. Append new rows to in-memory workbook
#   4. Write updated master.xlsx back to S3
#   If step 4 fails, the backup from step 2 is already safe.
# =============================================================================

def backup_and_append_excel(exports_bucket: str, items: list[dict],
                             receipt_meta: dict) -> bool:
    """
    Backup master.xlsx then append new rows.
    Returns True on success, False on failure (backup preserved either way).
    Called by processor Lambda after saving items to DynamoDB.

    items: list of item dicts from grocery-items table
    receipt_meta: dict with receipt_date, retailer_name, total_amount
    """
    import io
    import openpyxl

    s3     = boto3.client("s3")
    today  = datetime.now(timezone.utc).date().isoformat()
    master = "master.xlsx"
    backup = f"master_backup_{today}.xlsx"

    # Step 1: Read existing master.xlsx (or create fresh if first run)
    try:
        obj = s3.get_object(Bucket=exports_bucket, Key=master)
        wb  = openpyxl.load_workbook(io.BytesIO(obj["Body"].read()))
        ws  = wb.active
    except s3.exceptions.NoSuchKey:
        # First ever run â€” create fresh workbook with headers
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Receipts"
        ws.append(["Date", "Store", "Item (Raw)", "Item (Normalised)",
                   "Category", "Price", "Confidence", "Match Source", "Receipt ID"])

    # Step 2: Create dated backup BEFORE making any changes
    backup_buffer = io.BytesIO()
    wb.save(backup_buffer)
    backup_buffer.seek(0)
    try:
        s3.put_object(Bucket=exports_bucket, Key=backup,
                      Body=backup_buffer.read(),
                      ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        logger.info("Backup created: %s", backup)
    except Exception as e:
        logger.error("Failed to create backup %s: %s", backup, e)
        return False  # abort â€” don't write master if backup failed

    # Step 3: Append new rows (one row per line item)
    receipt_date    = receipt_meta.get("receipt_date", today)
    retailer_name   = receipt_meta.get("retailer_name", "Unknown")
    receipt_id      = receipt_meta.get("receipt_id", "")

    for item in items:
        ws.append([
            receipt_date,
            retailer_name,
            item.get("raw_name", ""),
            item.get("normalized_name", ""),
            item.get("category", ""),
            item.get("price", ""),
            item.get("confidence", ""),
            item.get("match_source", ""),
            receipt_id
        ])

    # Step 4: Write updated master.xlsx back to S3
    master_buffer = io.BytesIO()
    wb.save(master_buffer)
    master_buffer.seek(0)
    try:
        s3.put_object(Bucket=exports_bucket, Key=master,
                      Body=master_buffer.read(),
                      ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        logger.info("master.xlsx updated with %d new rows", len(items))
        return True
    except Exception as e:
        logger.error("Failed to write updated master.xlsx: %s", e)
        return False  # backup from step 2 is still safe


def list_backups(exports_bucket: str) -> list[dict]:
    """List all backup files available for restore, newest first."""
    s3   = boto3.client("s3")
    resp = s3.list_objects_v2(Bucket=exports_bucket, Prefix="master_backup_")
    files = resp.get("Contents", [])
    return sorted(
        [{"filename": f["Key"], "size": f["Size"],
          "last_modified": f["LastModified"].isoformat()} for f in files],
        key=lambda x: x["last_modified"],
        reverse=True
    )


def restore_backup(exports_bucket: str, backup_filename: str) -> bool:
    """Copy a backup file back as the current master.xlsx."""
    s3 = boto3.client("s3")
    try:
        s3.copy_object(
            Bucket=exports_bucket,
            CopySource={"Bucket": exports_bucket, "Key": backup_filename},
            Key="master.xlsx"
        )
        logger.info("Restored %s â†’ master.xlsx", backup_filename)
        return True
    except Exception as e:
        logger.error("Failed to restore backup %s: %s", backup_filename, e)
        return False


# =============================================================================
# REFERENCE DATA
# =============================================================================

def get_all_categories() -> list[dict]:
    resp = db().Table(CATEGORIES_TABLE).scan()
    return sorted(resp.get("Items", []), key=lambda x: x.get("name", ""))


def get_all_retailers() -> list[dict]:
    resp = db().Table(RETAILERS_TABLE).scan()
    return sorted(resp.get("Items", []), key=lambda x: x.get("name", ""))


def get_retailer(retailer_id: str) -> Optional[dict]:
    resp = db().Table(RETAILERS_TABLE).get_item(Key={"retailer_id": retailer_id})
    return resp.get("Item")


def save_retailer(retailer: dict):
    db().Table(RETAILERS_TABLE).put_item(Item=retailer)


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------
import io as _io
import re as _re_exp

MASTER_XLSX = "exports/master.xlsx"


def _parse_price(price_str):
    if not price_str:
        return 0.0
    cleaned = _re_exp.sub(r"[^\d.]", "", str(price_str))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def receipt_export_ready(receipt_id: str) -> bool:
    items = get_items(receipt_id)
    if not items:
        return False
    for item in items:
        needs_review = item.get("needs_review", False)
        confirmed    = item.get("confirmed", False)
        confidence   = float(item.get("match_confidence", 0))
        source       = item.get("match_source", "unknown")
        exact        = source in ("store_exact", "global_exact") and confidence >= 1.0
        if not exact and not confirmed:
            return False
    return True


def export_receipt_to_excel(receipt_id: str, exports_bucket: str):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    receipt = get_receipt(receipt_id)
    items   = get_items(receipt_id)
    if not receipt or not items:
        return None

    s3c = boto3.client("s3")

    try:
        obj = s3c.get_object(Bucket=exports_bucket, Key=MASTER_XLSX)
        wb  = load_workbook(_io.BytesIO(obj["Body"].read()))
        ws  = wb.active
        is_new = False
    except Exception:
        wb     = Workbook()
        ws     = wb.active
        ws.title = "Grocery Items"
        is_new = True

    HEADERS = [
        "Date", "Store", "Receipt ID",
        "Item Name", "Normalised Name",
        "Category", "Original Category",
        "Price (GBP)", "Quantity",
        "Confidence %", "Matched Keyword", "Match Source",
        "Corrected", "Corrected At", "Exported At"
    ]

    if is_new or ws.cell(1, 1).value != "Date":
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2E4057")
            cell.alignment = Alignment(horizontal="center")
        widths = [12,15,38,30,25,18,18,10,10,14,25,14,10,20,20]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    receipt_date = receipt.get("receipt_date", "")
    store_name   = receipt.get("retailer_name") or receipt.get("retailer_id", "")
    exported_at  = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")

    for item in items:
        conf_pct  = round(float(item.get("match_confidence", 0)) * 100, 1)
        corrected = "Yes" if item.get("confirmed") and item.get("match_source") == "manual" else "No"
        ws.append([
            receipt_date, store_name, receipt_id,
            item.get("raw_name", ""),
            item.get("normalized_name", ""),
            item.get("category", ""),
            item.get("original_category", ""),
            _parse_price(item.get("price", "0")),
            item.get("quantity", "1"),
            conf_pct,
            item.get("matched_keyword", ""),
            item.get("match_source", ""),
            corrected,
            item.get("corrected_at", ""),
            exported_at,
        ])

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3c.put_object(
        Bucket=exports_bucket,
        Key=MASTER_XLSX,
        Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    now = datetime.now(timezone.utc).isoformat()
    for item in items:
        db().Table(ITEMS_TABLE).update_item(
            Key={"receipt_id": receipt_id, "item_seq": item["item_seq"]},
            UpdateExpression="SET exported_to_excel = :t, exported_at = :ts, export_filename = :fn",
            ExpressionAttributeValues={":t": True, ":ts": now, ":fn": MASTER_XLSX}
        )

    logger.info(f"Exported {len(items)} items from {receipt_id} to {MASTER_XLSX}")
    return MASTER_XLSX
